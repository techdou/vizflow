"""Dataset parsing + schema inference utilities.

Supports CSV, JSON, and XLSX (if openpyxl is installed). Used at upload time to
persist the inferred schema (fix #8) so it isn't recomputed on every chart gen.
"""
import csv
import io
import json
import os
from typing import List, Dict, Any, Optional, Tuple


def _infer_col_type(values: List[str]) -> str:
    """Infer a Vega-Lite type for one column from a sample of string values."""
    non_empty = [v for v in values if v not in (None, "", "N/A")]
    if not non_empty:
        return "nominal"
    try:
        nums = [float(v) for v in non_empty]
        unique = len(set(nums))
        if unique <= 10 and all(n == int(n) for n in nums):
            return "ordinal"
        return "quantitative"
    except (ValueError, TypeError):
        sample = str(non_empty[0]).lower()
        if any(p in sample for p in ["date", "time", "year", "month", "-", "/"]):
            return "temporal"
        return "nominal"


def infer_schema(rows: List[dict], columns: List[str]) -> Dict[str, Any]:
    """Build a persisted schema: per-column inferred type + row count."""
    sample = rows[: min(50, len(rows))]
    fields = []
    for col in columns:
        col_vals = [r.get(col) for r in sample]
        fields.append({"name": col, "type": _infer_col_type(col_vals)})
    return {"fields": fields, "row_count": len(rows)}


def parse_csv(content: bytes) -> Tuple[List[str], List[dict]]:
    """Parse CSV bytes → (columns, rows-as-dicts)."""
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    rows = [dict(r) for r in reader]
    columns = list(rows[0].keys()) if rows else (reader.fieldnames or [])
    return columns, rows


def parse_json(content: bytes) -> Tuple[List[str], List[dict]]:
    """Parse JSON bytes → (columns, rows). Accepts a list of objects."""
    data = json.loads(content.decode("utf-8", errors="replace"))
    if isinstance(data, dict):
        # Allow {"data": [...]} or treat a single object as one row
        data = data.get("data", [data])
    if not isinstance(data, list):
        raise ValueError("JSON must be a list of objects")
    rows = [dict(r) for r in data if isinstance(r, dict)]
    columns = list(rows[0].keys()) if rows else []
    return columns, rows


def parse_xlsx(content: bytes) -> Tuple[List[str], List[dict]]:
    """Parse XLSX bytes → (columns, rows). Requires openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError("XLSX support requires openpyxl: pip install openpyxl") from e
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(h) if h is not None else f"col{i}" for i, h in enumerate(next(rows_iter))]
    except StopIteration:
        wb.close()
        return [], []
    rows = []
    for r in rows_iter:
        if all(c is None for c in r):
            continue
        rows.append({header[i]: ("" if r[i] is None else str(r[i])) for i in range(len(header))})
    wb.close()
    return header, rows


def detect_format(filename: str, content_type: Optional[str]) -> str:
    """Determine parser to use from filename extension + MIME (fix #6 fallback)."""
    name = (filename or "").lower()
    # Extension wins — MIME from browsers is unreliable for CSV.
    if name.endswith(".csv"):
        return "csv"
    if name.endswith(".json"):
        return "json"
    if name.endswith((".xlsx", ".xlsm")):
        return "xlsx"
    # Fall back to MIME if no recognized extension
    ct = (content_type or "").lower()
    if "csv" in ct or "text/plain" in ct:
        return "csv"
    if "json" in ct:
        return "json"
    if "spreadsheet" in ct or "excel" in ct:
        return "xlsx"
    raise ValueError(f"Unsupported file: '{filename}' (mime={content_type}). Supported: .csv .json .xlsx")


def parse_dataset(content: bytes, filename: str, content_type: Optional[str]) -> Tuple[List[str], List[dict]]:
    """Dispatch to the right parser based on detected format."""
    fmt = detect_format(filename, content_type)
    if fmt == "csv":
        return parse_csv(content)
    if fmt == "json":
        return parse_json(content)
    if fmt == "xlsx":
        return parse_xlsx(content)
    raise ValueError(f"Unsupported format: {fmt}")
